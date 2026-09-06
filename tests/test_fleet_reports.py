from datetime import date

from openpyxl import load_workbook

from services.fleet_reports import build_fleet_report, fleet_analytics, parse_maintenance_csv


def test_parse_motive_maintenance_preserves_manager_and_mxn():
    content = (
        "Date,Entity,Entity Type,Service Type,Service Name,Notes,Cost,Fleet Manager,Odometer,Engine Hours\n"
        "20/07/2026,87 ALFA AGS,vehicle,,,KIT CLUTCH,2900,SAUL PEREZ (AGS),200508,700\n"
    ).encode()
    rows = parse_maintenance_csv(content)
    assert len(rows) == 1
    assert rows[0]["amount_mxn"] == 2900
    assert rows[0]["submitted_by"] == "SAUL PEREZ (AGS)"
    assert rows[0]["vehicle_number"] == "87 ALFA AGS"


def test_empty_report_has_executive_sheets_without_misleading_empty_tabs(tmp_path):
    payload = build_fleet_report({"expenses": [], "driver_events": [], "speeding": [], "activity": [], "faults": []}, date(2026, 7, 1), date(2026, 7, 20))
    target = tmp_path / "report.xlsx"; target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True)
    assert workbook.sheetnames == ["Dashboard", "Resumen por chofer", "Unidades y diagnóstico"]


def test_office_expenses_sheet_has_an_explicit_name(tmp_path):
    payload = build_fleet_report({
        "expenses": [{
            "occurred_at": "2026-07-31", "zone_name": "Viaecos",
            "expense_type": "gasto_directo", "category": "Refacciones",
            "amount_mxn": 1200, "source": "ge_control_direct",
            "submitted_by": "Gastos y pagos",
        }],
        "driver_events": [], "speeding": [], "activity": [], "faults": [],
    }, date(2026, 7, 1), date(2026, 7, 31))
    target = tmp_path / "office-expenses.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True, data_only=True)
    assert "Gastos de oficina" in workbook.sheetnames
    assert "Gastos" not in workbook.sheetnames
    assert workbook["Gastos de oficina"]["C2"].value == "Viaecos"


def test_report_adds_visual_manager_dashboard_and_driver_summary(tmp_path):
    payload = build_fleet_report({
        "vehicles": [
            {"vehicle_number": "U-ROJA", "current_driver_name": "Ana"},
            {"vehicle_number": "U-SIN-GPS", "current_driver_name": "Luis"},
        ],
        "driver_events": [
            {"vehicle_number": "U-ROJA", "driver_name": "Ana", "primary_behavior": "hard_brake", "severity": "high"}
            for _ in range(3)
        ],
        "speeding": [],
        "faults": [
            {"vehicle_number": "U-ROJA", "code": "P0201", "occurrence_count": 4},
            {"vehicle_number": "U-ROJA", "code": "P0201", "occurrence_count": 2},
            {"vehicle_number": "U-ROJA", "code": "P0351", "occurrence_count": 1},
        ],
        "inspections": [{"vehicle_number": "U-ROJA", "driver_name": "Ana"}],
        "expenses": [], "activity": [],
    }, date(2026, 7, 1), date(2026, 7, 7), "Zacatecas")
    target = tmp_path / "visual-report.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, data_only=True)

    dashboard = workbook["Dashboard"]
    dashboard_values = [cell.value for row in dashboard.iter_rows() for cell in row]
    assert "INFORME EJECUTIVO · FLOTILLA 360" in dashboard_values
    assert "Choferes que requieren capacitación" in dashboard_values
    assert "Unidades sin datos GPS" in dashboard_values
    assert "UNIDADES SIN DATOS GPS" in dashboard_values
    assert "Choferes que realizaron inspecciones" in dashboard_values
    assert "DECISIONES RECOMENDADAS PARA EL GERENTE" in dashboard_values
    assert "Ana" in dashboard_values
    assert any("Capacitación en distancia y frenado preventivo" in str(value) for value in dashboard_values)
    assert dashboard["O2"].value == "Frenado brusco"
    assert dashboard["P2"].value == 3
    assert len(dashboard._charts) == 1
    assert dashboard._charts[0].__class__.__name__ == "DoughnutChart"

    driver_summary = workbook["Resumen por chofer"]
    assert [cell.value for cell in driver_summary[2]] == [
        "Chofer", "Eventos de seguridad", "Excesos de velocidad", "Eventos totales",
        "Conducta principal", "Críticos / altos", "Inspecciones", "Capacitación recomendada", "Prioridad",
    ]
    assert driver_summary["A3"].value == "Ana"
    assert driver_summary["E3"].value == "Frenado brusco"


def test_report_adds_executive_chronology_when_events_exist(tmp_path):
    payload = build_fleet_report({
        "driver_events": [{
            "started_at": "2026-07-10T12:00:00Z", "vehicle_number": "U-1",
            "driver_name": "Ana", "primary_behavior": "hard_brake", "severity": "high",
        }],
        "speeding": [], "faults": [], "expenses": [], "activity": [],
    }, date(2026, 7, 1), date(2026, 7, 20))
    target = tmp_path / "report.xlsx"; target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True)
    assert "Eventos de chofer" in workbook.sheetnames
    sheet = workbook["Eventos de chofer"]
    assert sheet["E2"].value == "Frenado brusco"


def test_report_converts_utc_timestamps_to_mexico_local_time(tmp_path):
    payload = build_fleet_report({
        "inspections": [{
            "inspected_at": "2026-08-01T00:12:53+00:00", "driver_name": "Ana",
            "vehicle_number": "U-1", "inspection_type": "post_trip", "status": "resolved",
        }],
        "driver_events": [], "speeding": [], "faults": [], "expenses": [], "activity": [],
    }, date(2026, 7, 31), date(2026, 8, 1))
    target = tmp_path / "local-time-report.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, data_only=True)
    assert workbook["Inspecciones"]["A2"].value.isoformat() == "2026-07-31T18:12:53"


def test_analytics_uses_consumed_fuel_and_exact_utilization_rollup():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1"}],
        "metrics": [{"vehicle_number": "U-1", "metric_date": "2026-07-01", "distance_km": 100}],
        "fuel": [{"vehicle_number": "U-1", "quantity_liters": 30, "total_cost": 900}],
        "utilization": [{
            "vehicle_number": "U-1", "engine_hours": 8, "fuel_consumed_liters": 20,
            "utilization_pct": 0.8,
        }],
        "_sync": {"datasets": {"card_expenses": {"status": "unavailable"}}},
    })
    unit = analytics["units"][0]
    assert unit["purchased_liters"] == 30
    assert unit["liters"] == 20
    assert unit["engine_hours"] == 8
    assert unit["utilization_pct"] == 0.8
    assert unit["km_per_liter"] == 5
    assert analytics["totals"]["expense_complete"] is False


def test_analytics_prioritizes_exact_mileage_rollup_without_double_counting():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1"}],
        "mileage": [{"vehicle_number": "U-1", "distance_km": 250}],
        "metrics": [{"vehicle_number": "U-1", "metric_date": "2026-07-01", "distance_km": 100}],
        "activity": [{"vehicle_number": "U-1", "started_at": "2026-07-01", "distance_km": 75}],
    })
    assert analytics["units"][0]["distance_km"] == 250
    assert analytics["totals"]["distance_km"] == 250


def test_analytics_ranks_by_event_count_and_recovers_driver_from_activity():
    analytics = fleet_analytics({
        "vehicles": [
            {"vehicle_number": "U-1", "current_driver_name": ""},
            {"vehicle_number": "U-2", "current_driver_name": ""},
            {"vehicle_number": "SIN-GPS", "current_driver_name": ""},
        ],
        "driver_events": [
            {"vehicle_number": "U-1", "severity": "low", "primary_behavior": "hard_brake"},
            {"vehicle_number": "U-1", "severity": "low", "primary_behavior": "hard_brake"},
            {"vehicle_number": "U-2", "severity": "critical", "primary_behavior": "cell_phone"},
        ],
        "speeding": [{"vehicle_number": "U-1", "severity": "medium"}],
        "activity": [{"vehicle_number": "U-1", "driver_name": "CONDUCTOR REAL", "started_at": "2026-07-20"}],
        "_period_days": 7,
    })

    assert [row["vehicle_number"] for row in analytics["units"]] == ["U-1", "U-2", "SIN-GPS"]
    assert analytics["units"][0]["driver_name"] == "CONDUCTOR REAL"
    assert analytics["behaviors"][0] == {"label": "Frenado brusco", "count": 2}
    assert {"label": "Exceso de velocidad", "count": 1} in analytics["behaviors"]
    assert analytics["totals"]["vehicles_with_data"] == 2
    assert analytics["totals"]["vehicles_without_gps"] == 1
    assert analytics["units"][-1]["coverage_status"] == "Sin datos GPS / revisión manual"
    assert [row["vehicle_number"] for row in analytics["attention_units"]] == ["U-1", "U-2"]
    assert [row["vehicle_number"] for row in analytics["units_without_gps"]] == ["SIN-GPS"]


def test_analytics_attributes_inspections_to_reported_driver_with_unit_fallback():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Chofer asignado"}],
        "inspections": [
            {"vehicle_number": "U-1", "driver_name": "Quien inspeccionó"},
            {"vehicle_number": "U-1", "driver_name": ""},
        ],
    })

    assert analytics["inspection_credits"] == [
        {"vehicle_number": "U-1", "driver_name": "Chofer asignado", "inspections": 1},
        {"vehicle_number": "U-1", "driver_name": "Quien inspeccionó", "inspections": 1},
    ]


def test_inspection_driver_without_events_is_visible_and_explains_missing_assignment():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "UTILITARIA MIGUEL", "current_driver_name": ""}],
        "inspections": [{"vehicle_number": "UTILITARIA MIGUEL", "driver_name": "MIGUEL ANGEL"}],
        "driver_events": [], "speeding": [],
    })

    assert analytics["units_without_gps"] == []
    assert analytics["units"][0]["coverage_status"] == "Con datos GPS"
    assert analytics["drivers_without_events"] == [{
        "driver_name": "MIGUEL ANGEL",
        "vehicle_number": "UTILITARIA MIGUEL",
        "inspections": 1,
    }]


def test_out_of_service_unit_is_not_reported_as_missing_gps_activity():
    analytics = fleet_analytics({
        "vehicles": [
            {"vehicle_number": "TALLER", "availability_status": "out_of_service"},
            {"vehicle_number": "OPERATIVA", "status": "active"},
        ],
    })

    assert [row["vehicle_number"] for row in analytics["units_without_gps"]] == ["OPERATIVA"]
    assert analytics["totals"]["vehicles_without_gps"] == 1


def test_inspection_dashboard_separates_open_pending_from_total_completed_work():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Ana"}],
        "inspections": [
            {"id": 11, "vehicle_number": "U-1", "driver_name": "Ana"},
            {"id": 12, "vehicle_number": "U-1", "driver_name": "Ana"},
        ],
        "defects": [
            {"inspection_id": 11, "vehicle_number": "U-1", "status": "open"},
            {"inspection_id": 11, "vehicle_number": "U-1", "status": "open"},
            {"inspection_id": 12, "vehicle_number": "U-1", "status": "resolved", "resolved_at": "2026-08-27"},
        ],
    })

    assert analytics["inspection_credits"] == [{"vehicle_number": "U-1", "driver_name": "Ana", "inspections": 2}]
    assert analytics["pending_inspection_credits"] == [{"vehicle_number": "U-1", "driver_name": "Ana", "inspections": 1}]


def test_expense_units_keep_the_unit_and_assigned_driver_for_dashboard():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Ana"}],
        "fuel": [{"vehicle_number": "U-1", "quantity_liters": 42, "total_cost": 900, "currency": "MXN"}],
        "_sync": {"datasets": {"card_expenses": {"status": "unavailable"}}},
    })

    assert analytics["expense_units"] == [{
        "vehicle_number": "U-1", "driver_name": "Ana", "expenses_mxn": 900.0, "purchased_liters": 42.0,
    }]


def test_general_expenses_are_not_reported_as_a_unit_without_gps():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Ana"}],
        "expenses": [{"vehicle_number": "", "amount_mxn": 1200, "source": "ge_control_direct"}],
    })

    assert analytics["expense_units"] == []
    assert analytics["general_expenses"] == {
        "expenses_mxn": 1200.0, "purchased_liters": 0.0,
        "records": 1, "direct_invoices": 1,
        "details": [{
            "date": None, "invoice_number": "Sin folio",
            "supplier": "Proveedor no identificado", "concept": "Sin concepto",
            "description": "", "amount_mxn": 1200.0,
        }],
    }
    assert all(row["vehicle_number"] != "Sin unidad vinculada" for row in analytics["units_without_gps"])
    assert analytics["totals"]["vehicles_without_gps"] == 1


def test_missing_inspections_uses_last_seen_driver_without_claiming_current_assignment():
    analytics = fleet_analytics({
        "vehicles": [{
            "vehicle_number": "C 75", "current_driver_name": "",
            "last_known_driver_name": "RUBEN MORENO",
            "last_known_driver_at": "2026-08-11T19:44:03Z",
        }],
        "inspections": [],
    })

    assert analytics["units_without_inspections"] == [{
        "vehicle_number": "C 75", "driver_name": "RUBEN MORENO",
        "driver_context": "Último chofer visto",
        "driver_last_seen_at": "2026-08-11T19:44:03Z",
    }]


def test_analytics_attributes_training_to_event_driver_not_only_assigned_unit_driver():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "U-1", "current_driver_name": "Chofer asignado"}],
        "driver_events": [{
            "vehicle_number": "U-1", "driver_name": "Chofer real",
            "primary_behavior": "hard_brake", "severity": "high",
        }],
        "speeding": [],
        "inspections": [{"vehicle_number": "U-1", "driver_name": "Chofer real"}],
    })

    driver = analytics["training_drivers"][0]
    assert driver["driver_name"] == "Chofer real"
    assert driver["top_behavior"] == "Frenado brusco"
    assert driver["training"] == "Capacitación en distancia y frenado preventivo"
    assert driver["inspections"] == 1


def test_excel_keeps_units_out_of_driver_dashboard_and_in_unit_sheet(tmp_path):
    vehicles = [{"vehicle_number": f"U-{index:02d}"} for index in range(1, 14)]
    events = [
        {"vehicle_number": row["vehicle_number"], "primary_behavior": "hard_brake", "severity": "low"}
        for row in vehicles
    ]
    payload = build_fleet_report(
        {"vehicles": vehicles, "driver_events": events, "speeding": [], "faults": [], "activity": []},
        date(2026, 7, 20),
        date(2026, 7, 25),
    )
    target = tmp_path / "report.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True, data_only=True)

    dashboard = workbook["Dashboard"]
    dashboard_values = [cell.value for row in dashboard.iter_rows() for cell in row]
    summary_headers = [cell.value for cell in workbook["Unidades y diagnóstico"][2]]
    assert not any("U-13" in str(value) for value in dashboard_values)
    unit_values = [cell.value for row in workbook["Unidades y diagnóstico"].iter_rows() for cell in row]
    assert "U-13" in unit_values
    assert "Índice" not in dashboard_values
    assert "SCORE CONDUCTORES" not in dashboard_values
    assert "Score" not in summary_headers
    assert "Índice de atención" not in summary_headers
    assert "Actividad" not in workbook.sheetnames
    assert "Scorecard conductores" not in workbook.sheetnames


def test_foreign_currency_fuel_is_not_reported_as_mxn_expense():
    analytics = fleet_analytics({
        "vehicles": [{"vehicle_number": "AT 35"}],
        "fuel": [{
            "vehicle_number": "AT 35",
            "quantity_liters": 187.95,
            "total_cost": 23.99,
            "currency": "CAD",
        }],
        "expenses": [],
        "_sync": {"datasets": {"card_expenses": {"status": "unavailable"}}},
    })

    assert analytics["totals"]["expenses_mxn"] == 0
    assert analytics["totals"]["expense_available"] is False


def test_closed_faults_and_resolved_defects_are_excluded_from_excel_alerts(tmp_path):
    payload = build_fleet_report({
        "vehicles": [{"vehicle_number": "U-1"}],
        "faults": [
            {"vehicle_number": "U-1", "code": "P-OPEN", "occurrence_count": 2, "status": "open"},
            {"vehicle_number": "U-1", "code": "P-CLOSED", "occurrence_count": 99, "status": "closed"},
        ],
        "defects": [
            {"vehicle_number": "U-1", "title": "Freno abierto", "severity": "major", "status": "open"},
            {"vehicle_number": "U-1", "title": "Freno reparado", "severity": "critical", "status": "resolved", "resolved_at": "2026-08-01"},
        ],
        "driver_events": [], "speeding": [], "expenses": [], "activity": [],
    }, date(2026, 7, 20), date(2026, 8, 4))
    target = tmp_path / "open-only.xlsx"
    target.write_bytes(payload)
    workbook = load_workbook(target, read_only=True, data_only=True)

    assert [row[2] for row in workbook["Diagnóstico PID"].iter_rows(min_row=2, values_only=True)] == ["P-OPEN"]
    assert [row[2] for row in workbook["Defectos accionables"].iter_rows(min_row=2, values_only=True)] == ["Freno abierto"]
    unit_row = next(workbook["Unidades y diagnóstico"].iter_rows(min_row=3, values_only=True))
    assert unit_row[7] == 2
    assert unit_row[17] == 1
